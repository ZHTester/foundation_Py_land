# encoding: utf-8

"""
# @Time    : 17/3/2020 10:20 下午
# @Author  : Function
# @FileName    : use_Excel.py
# @Software: PyCharm
操作Excel
openpyxl 操作Excel文件路径
"""
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image  # 插入图片
from openpyxl.styles import Font, colors  # 插入文字大小  和  颜色
import pymysql


class ExcelUtil:
    def __init__(self):
        self.wb = Workbook()  # 创建Excel文件对象
        self.ws = self.wb.active  # 默认激活表单
        self.ws.title = '这是我的表单'
        # 创建表单
        self.ws_two = self.wb.create_sheet("我创建的表单")
        self.ws_three = self.wb.create_sheet()  # 默认创建表单

    def do_save(self):
        # 插入文字
        self.ws['A2'] = "这是我插入的文件数据"
        self.ws['A1'] = "这是我插入的文件数据1"
        # 设置字体颜色
        font = Font(sz=55, color=colors.RED)
        self.ws['A2'].font = font
        # 循环插入
        for row in self.ws_three["A1:E5"]:
            for cell in row:
                cell.value = 2
        # 数据插入
        self.ws_three['G1'] = '=SUM(A1:E1)'

        # 插入图片
        img = Image('./static/123.jpg')

        # 合并单元格
        self.ws.merge_cells('A4:E4')
        self.ws.unmerge_cells('A4:E4')
        self.ws.add_image(img, 'B1')
        self.wb.save('./static/我创建的.xlsx')

    def do_reade(self):
        """
        读取文件内容
        :return:
        """
        wb = load_workbook('./static/data.xlsx')  # 载入文件
        names = wb.get_sheet_names()  # 输出excel文件的sheet_name
        # 获取每一个单元格的数据
        wbb = wb[names[1]]
        for row in wbb.rows:
            for cell in row:
                if cell.value is None:
                    pass
                else:
                    print(cell.value)

    def insert_mysql(self):
        """
        Excel  《-----------------》  插入数据库
        :return:
        """
        wb = load_workbook('./static/data.xlsx')  # 载入文件
        names = wb.get_sheet_names()  # 输出excel文件的sheet_name

        # 查询数据
        wb1 = wb[names[1]]
        for i in range(wb1.max_row):
            if i == 0:
                continue
            user_name1 = wb1['A{0}'.format(i)].value
            over1 = wb1['B{0}'.format(i)].value
            # Excel数据<---->写入数据
            conn = self.get_conn()
            cursor = conn.cursor()  # 加入游标
            sql1 = 'INSERT INTO user1(user_name,over) VALUES("{0}","{1}")'.format(user_name1, over1)   # 执行sql
            cursor.execute(sql1)
        print('。。。。。执行成功。。。。。。')


    def write_mysql(self):
        xueli = []
        city = []
        salary = []

        cc = 0
        bb = 0
        dd = 0
        ee = 0
        ff = 0

        conn = self.get_conn()
        cursor = conn.cursor()  # 加入游标
        sql1 = """select degree_need,job_city,salary from lagou_job"""
        cursor.execute(sql1)
        # print(cursor.execute(sql1))
        results = cursor.fetchall()
        for (i, value) in enumerate(results):
            value = list(value)
            for va in value:
                if '北京' in va:
                    xueli.append(value[0])
                    bb += 1
                    if '本科及以上' in value[0]:
                        cc += 1
                    elif '硕士及以上' in value[0]:
                        ee = 1
                    elif '学历不限' in value[0]:
                        ff += 1
                    elif '专科及以上' in value[0]:
                        dd += 1
        print("在北京的职位一共为{0}个职位".format(bb))
        print("本科职位为{0}个职位".format(cc))
        print("硕士职位为{0}个职位".format(ee))
        print("学历不限职位为{0}个职位".format(ff))

        #     xueli.append(value[0])
        #     city.append(value[1])
        #     salary.append(value[2])
        # print(list(set(xueli)))
        # print(list(set(city)))
        # print(sorted(list(set(salary))))




    def get_conn(self):
        """
        获取mysql的链接
        :return:
        """
        conn = pymysql.connect(
            # db='StudyTest',
            db='cnblogs_spider',
            host='127.0.0.1',
            user='root',
            password='admin123'
        )
        return conn


if __name__ == '__main__':
    c = ExcelUtil()
    c.write_mysql()

